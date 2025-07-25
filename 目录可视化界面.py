# 导入必要的库
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from datetime import datetime
import logging
import tkinter as tk
from tkinter import filedialog, messagebox


# 函数：获取指定目录的结构信息，支持四级文件夹层级，并合并相同文件夹名称
# 参数：
#   - root_dir: 要获取结构的根目录路径
# 返回值：包含目录结构信息的列表

def get_directory_structure(root_dir):
    """获取目录结构信息，支持四级文件夹层级，合并相同文件夹名称"""
    structure = []
    # 获取一级文件夹（上级文件夹）名称
    parent_dir = os.path.basename(root_dir)
    
    # 遍历当前目录下的所有内容（文件和文件夹），并按文件夹优先、名称排序
    items = os.listdir(root_dir)
    items.sort(key=lambda x: (not os.path.isdir(os.path.join(root_dir, x)), x))
    
    # 遍历每个项目，处理文件和文件夹
    for item in items:
        item_path = os.path.join(root_dir, item)
        
        # 处理二级文件夹
        if os.path.isdir(item_path):
            level2 = item
            
            # 遍历二级文件夹下的内容
            level3_items = os.listdir(item_path)
            level3_items.sort(key=lambda x: (not os.path.isdir(os.path.join(item_path, x)), x))
            
            # 检查二级文件夹下是否有内容
            if not level3_items:
                # 二级文件夹为空
                structure.append({
                    '一级文件夹': parent_dir,
                    '二级文件夹': level2,
                    '三级文件夹': "无",
                    '四级文件夹': "无",
                    '文件名': "无",
                    '格式': "文件夹"
                })
            else:
                for level3_item in level3_items:
                    level3_path = os.path.join(item_path, level3_item)
                    
                    if os.path.isdir(level3_path):
                        level3 = level3_item
                        
                        # 处理三级文件夹下的内容
                        level4_items = os.listdir(level3_path)
                        level4_items.sort(key=lambda x: (not os.path.isdir(os.path.join(level3_path, x)), x))
                        
                        if not level4_items:
                            # 三级文件夹为空
                            structure.append({
                                '一级文件夹': parent_dir,
                                '二级文件夹': level2,
                                '三级文件夹': level3,
                                '四级文件夹': "无",
                                '文件名': "无",
                                '格式': "文件夹"
                            })
                        else:
                            for level4_item in level4_items:
                                level4_path = os.path.join(level3_path, level4_item)
                                
                                if os.path.isdir(level4_path):
                                    # 四级文件夹
                                    level4 = level4_item
                                    
                                    # 处理四级文件夹下的内容
                                    level5_items = os.listdir(level4_path)
                                    level5_items.sort(key=lambda x: (not os.path.isdir(os.path.join(level4_path, x)), x))
                                    
                                    if not level5_items:
                                        # 四级文件夹为空
                                        structure.append({
                                            '一级文件夹': parent_dir,
                                            '二级文件夹': level2,
                                            '三级文件夹': level3,
                                            '四级文件夹': level4,
                                            '文件名': "无",
                                            '格式': "文件夹"
                                        })
                                    else:
                                        for level5_item in level5_items:
                                            # 处理四级文件夹下的文件或文件夹
                                            if os.path.isdir(os.path.join(level4_path, level5_item)):
                                                # 五级文件夹（超出四级，按文件处理）
                                                file_name, file_ext = os.path.splitext(level5_item)
                                                structure.append({
                                                    '一级文件夹': parent_dir,
                                                    '二级文件夹': level2,
                                                    '三级文件夹': level3,
                                                    '四级文件夹': level4,
                                                    '文件名': file_name,
                                                    '格式': "文件夹"
                                                })
                                            else:
                                                # 四级文件夹下的文件
                                                file_name, file_ext = os.path.splitext(level5_item)
                                                structure.append({
                                                    '一级文件夹': parent_dir,
                                                    '二级文件夹': level2,
                                                    '三级文件夹': level3,
                                                    '四级文件夹': level4,
                                                    '文件名': file_name,
                                                    '格式': file_ext[1:] if file_ext else "无格式"
                                                })
                                else:
                                    # 三级文件夹下的文件
                                    file_name, file_ext = os.path.splitext(level4_item)
                                    structure.append({
                                        '一级文件夹': parent_dir,
                                        '二级文件夹': level2,
                                        '三级文件夹': level3,
                                        '四级文件夹': "无",
                                        '文件名': file_name,
                                        '格式': file_ext[1:] if file_ext else "无格式"
                                    })
                    else:
                        # 二级文件夹下的文件
                        file_name, file_ext = os.path.splitext(level3_item)
                        structure.append({
                            '一级文件夹': parent_dir,
                            '二级文件夹': level2,
                            '三级文件夹': "无",
                            '四级文件夹': "无",
                            '文件名': file_name,
                            '格式': file_ext[1:] if file_ext else "无格式"
                        })
        else:
            # 根目录下的文件
            file_name, file_ext = os.path.splitext(item)
            structure.append({
                '一级文件夹': parent_dir,
                '二级文件夹': "无",
                '三级文件夹': "无",
                '四级文件夹': "无",
                '文件名': file_name,
                '格式': file_ext[1:] if file_ext else "无格式"
            })
    
    return structure

# 函数：根据目录结构信息创建并美化 Excel 文件，合并相同文件夹名称的单元格
# 参数：
#   - structure: 包含目录结构信息的列表
#   - output_file: 输出的 Excel 文件路径

def create_excel(structure, output_file):
    """创建并美化Excel文件，合并相同文件夹名称的单元格"""
    # 创建一个新的工作簿，并获取活动工作表
    wb = Workbook()
    ws = wb.active
    ws.title = "目录清单"
    
    # 更新表头
    headers = ["一级文件夹", "二级文件夹", "三级文件夹", "四级文件夹", "文件名", "格式"]
    ws.append(headers)
    
    # 添加数据
    for row in structure:
        ws.append([
            row["一级文件夹"], 
            row["二级文件夹"], 
            row["三级文件夹"], 
            row["四级文件夹"], 
            row["文件名"], 
            row["格式"]
        ])
    
    # 合并相同内容的单元格（按列处理）
    merge_columns = ['A', 'B', 'C', 'D']  # 一到四级文件夹列
    for col_letter in merge_columns:
        current_value = None
        start_row = 2
        for row in range(2, ws.max_row + 1):
            cell_value = ws[f'{col_letter}{row}'].value
            if current_value is None:
                current_value = cell_value
            elif cell_value != current_value:
                if row - 1 > start_row:
                    ws.merge_cells(f'{col_letter}{start_row}:{col_letter}{row-1}')
                current_value = cell_value
                start_row = row
        # 处理最后一组
        if ws.max_row > start_row:
            ws.merge_cells(f'{col_letter}{start_row}:{col_letter}{ws.max_row}')
    
    # 设置样式
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))
    content_alignment = Alignment(vertical="center")
    
    # 应用表头样式
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = header_fill
        cell.border = thin_border
    # 设置表头行高
    ws.row_dimensions[1].height = 20

    # 应用内容样式
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", horizontal="general", wrap_text=True)
    
    # 前4列单元格自动居中
    for col in range(1, 5):
        for row in ws.iter_rows(min_row=2):
            cell = row[col-1]
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                # 考虑换行后的内容长度
                if cell.value:
                    lines = str(cell.value).split('\n')
                    for line in lines:
                        if len(line) > max_length:
                            max_length = len(line)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 自动调整行高以显示全部内容
    for row in ws.iter_rows():
        max_height = 0
        for cell in row:
            try:
                # 根据内容行数和字体大小计算行高
                lines = str(cell.value).count('\n') + 1
                height = lines * 15
                if height > max_height:
                    max_height = height
            except:
                pass
        ws.row_dimensions[row[0].row].height = max_height

    wb.save(output_file)

# 函数：打开目录选择对话框，将选择的目录路径显示在输入框中

def select_directory():
    # 打开目录选择对话框
    directory = filedialog.askdirectory()
    if directory:
        # 清空输入框并插入选择的目录路径
        dir_entry.delete(0, tk.END)
        dir_entry.insert(0, directory)

# 函数：根据输入目录生成 Excel 文件，自动命名输出文件

def generate_excel():
    # 获取输入框中的目录路径
    input_dir = dir_entry.get()
    if not input_dir:
        # 若未选择目录，显示错误提示
        messagebox.showerror("错误", "请选择输入目录")
        return

    try:
        # 获取指定目录的结构信息
        structure = get_directory_structure(input_dir)
        if structure:
            # 从第一条记录中获取一级文件夹名称
            parent_dir = structure[0]['一级文件夹']
            # 按规则生成输出文件路径，将文件放在选择的目录下
            output_file = os.path.join(input_dir, f'{parent_dir}目录.xlsx')
        else:
            # 若未获取到结构信息，使用默认文件名，将文件放在选择的目录下
            output_file = os.path.join(input_dir, '未知目录.xlsx')
        # 根据结构信息创建 Excel 文件
        create_excel(structure, output_file)
        # 显示成功提示，包含生成的文件名
        messagebox.showinfo("成功", f"Excel 文件 {os.path.basename(output_file)} 生成成功！")
    except Exception as e:
        # 若生成过程中出错，显示错误信息
        messagebox.showerror("错误", f"生成失败：{str(e)}")


# 创建主窗口
root = tk.Tk()
# 设置窗口标题
root.title("目录结构可视化工具")

# 添加标题文字
title_label = tk.Label(root, text="技术管理中心归档目录生成工具", font=('Arial', 14))
title_label.pack(pady=10)

# 设置窗口大小
root.geometry("600x200")

# 输入目录选择
frame_dir = tk.Frame(root, padx=10, pady=10)
frame_dir.pack(fill=tk.X)

dir_label = tk.Label(frame_dir, text="选择输入目录：")
dir_label.pack(side=tk.LEFT)

dir_entry = tk.Entry(frame_dir, width=50)
dir_entry.pack(side=tk.LEFT, padx=5)

dir_button = tk.Button(frame_dir, text="浏览", command=select_directory)
dir_button.pack(side=tk.LEFT)

# 生成按钮
generate_button = tk.Button(root, text="生成 Excel", command=generate_excel, padx=20, pady=10)
generate_button.pack(pady=20)

# 运行 Tkinter 主循环，使窗口保持显示状态
root.mainloop()